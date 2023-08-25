import jira.client
from jira.client import JIRA
from openpyxl import Workbook

jira_options={'server': 'https://rchase.atlassian.net'}
jira=JIRA(options=jira_options,basic_auth=('firdous.shaikh@capgemini.com','ATATT3xFfGF0E1EoUI8oBgg6Fntsi9UBjKFvcPgYPk8HjR_Ixu1kzzpXKkOD4PO6vR1YVYZDS10bjVpOZgqwyLj1fY2qIfxrCgBJ-xsXfACXS0LqpLgm5Wx9qATKDX76FRXYpYTVFmb4BOP7xgdt_AVVeaawRSdfWT29eJHYHYBYNXYC6dtHEsw=DB5CCF3A'))



key_list = []
summary_list = []

#Add additional lists for fields here
#Example: 
#description_list = []



issues_in_project = jira.search_issues('project=RSRE')

for issue in issues_in_project:
    key_list.append(issue.key)
    summary_list.append(issue.fields.summary)

# Add additional fields returned here
# Example:
#. description_list.append(issue.fields.description)


wb = Workbook()
ws = wb.active
key_row = 1
summary_row = 1
#add additional "$FIELD_row = 1" entries here so the field results start at row 1
#Example:
#description_row = 1


start_column = 1



for key in key_list:
 ws.cell(row=key_row, column=start_column).value = key
 key_row += 1

for summary in summary_list:
 ws.cell(row=summary_row, column=start_column+1).value = summary
 summary_row += 1

# add additional fields here 
#Example:
#for description in description_list:
 # ws.cell(row=description_row, column=start_column+2).value = description
 # description_row += 1
wb.template=False
wb.save("jira-report.xlsx")